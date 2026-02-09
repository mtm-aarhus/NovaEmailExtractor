from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
SharepointUrl = orchestrator_connection.get_constant('AarhusKommuneSharepoint').value

def sharepoint_client(tenant: str, client_id: str, thumbprint: str, cert_path: str, sharepoint_site_url: str, orchestrator_connection: OrchestratorConnection) -> ClientContext:
    """
    Creates and returns a SharePoint client context.
    """
    # Authenticate to SharePoint
    cert_credentials = {
        "tenant": tenant,
        "client_id": client_id,
        "thumbprint": thumbprint,
        "cert_path": cert_path
    }
    ctx = ClientContext(sharepoint_site_url).with_client_certificate(**cert_credentials)

    # Load and verify connection
    web = ctx.web
    ctx.load(web)
    ctx.execute_query()

    orchestrator_connection.log_info(f"Authenticated successfully. Site Title: {web.properties['Title']}")
    return ctx

def upload_to_sharepoint(client: ClientContext, folder_name: str, file_path: str, folder_url: str):
    """
    Uploads a file to a specific folder in a SharePoint document library.

    :param client: Authenticated SharePoint client context
    :param folder_name: Name of the target folder within the document library
    :param file_path: Local file path to upload
    :param folder_url: SharePoint folder URL where the file should be uploaded
    """
    try:
        # Extract file name safely
        file_name = os.path.basename(file_path)

        # Define the SharePoint document library structure
        document_library = f"{folder_url.split('/', 1)[-1]}"
        folder_path = f"{document_library}/{folder_name}"

        # Read file into memory (Prevents closed file issue)
        with open(file_path, "rb") as file:
            file_content = file.read()  

        # Get SharePoint folder reference
        target_folder = client.web.get_folder_by_server_relative_url(folder_url)

        # Upload file using byte content
        target_folder.upload_file(file_name, file_content)
        
        # Execute request
        client.execute_query()
        orchestrator_connection.log_info(f"✅ Successfully uploaded: {file_name} to {folder_path}")

    except Exception as e:
        orchestrator_connection.log_info(f"❌ Error uploading file: {str(e)}")



# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Indsender emails"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Headers
headers = ["Case number", "Email"]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
for row_idx, item in enumerate(all_emails, start=2):
    ws.cell(row=row_idx, column=1, value=item["CaseNumber"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=item["Email"]).border = thin_border

# Column widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 35

# Freeze header row
ws.freeze_panes = "A2"

# Save file
output_path = "Indsender_emails.xlsx"
wb.save(output_path)

print(f"Excel file created: {output_path}")

orchestrator_connection.log_info('Overfører excelfil til sharepoint')
file_url = f'{SharepointUrl}/Teams/sec-lukket1752/Delte Dokumenter'

certification = orchestrator_connection.get_credential("SharePointCert")
api = orchestrator_connection.get_credential("SharePointAPI")

tenant = api.username
client_id = api.password
thumbprint = certification.username
cert_path = certification.password

client = sharepoint_client(tenant, client_id, thumbprint, cert_path, f'{SharepointUrl}/Teams/tea-teamsite11160/', orchestrator_connection)

upload_to_sharepoint(client= client, folder_name = 'Delte Dokumenter', file_path=output_path, folder_url= '/Teams/tea-teamsite11160/Delte Dokumenter')
orchestrator_connection.log_info(f'Uploaded to {file_url}')