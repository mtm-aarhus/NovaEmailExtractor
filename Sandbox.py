from GetKmdAcessToken import GetKMDToken
import requests
import os
import uuid
from datetime import datetime, timedelta
import json
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
orchestrator_connection = OrchestratorConnection("AktbobGenererAktindsigter", os.getenv('OpenOrchestratorSQL'),os.getenv('OpenOrchestratorKey'), None,None)
# ---- Henter assests og credentials -----
KMDNovaURL = orchestrator_connection.get_constant("KMDNovaURL").value
SharepointUrl = orchestrator_connection.get_constant('AarhusKommuneSharepoint').value
SharepointUrl = orchestrator_connection.get_constant('AarhusKommuneSharepoint').value

  # ---- Henter access tokens ----
KMD_access_token = GetKMDToken(orchestrator_connection)


def sharepoint_client(tenant: str, client_id: str, thumbprint: str, cert_path: str, sharepoint_site_url: str, orchestrator_connection: OrchestratorConnection) -> ClientContext:
        """
        Creates and returns a SharePoint client context.
        """
        # Authenticate to SharePoint
        cert_credentials = {
            "tenant": tenant,
            "client_id": client_id,
            "thumbprint": thumbprint,
            "cert_path": cert_path
        }
        ctx = ClientContext(sharepoint_site_url).with_client_certificate(**cert_credentials)

        # Load and verify connection
        web = ctx.web
        ctx.load(web)
        ctx.execute_query()

        orchestrator_connection.log_info(f"Authenticated successfully. Site Title: {web.properties['Title']}")
        return ctx

def upload_to_sharepoint(client: ClientContext, folder_name: str, file_path: str, folder_url: str):
            """
            Uploads a file to a specific folder in a SharePoint document library.

            :param client: Authenticated SharePoint client context
            :param folder_name: Name of the target folder within the document library
            :param file_path: Local file path to upload
            :param folder_url: SharePoint folder URL where the file should be uploaded
            """
            try:
                # Extract file name safely
                file_name = os.path.basename(file_path)

                # Define the SharePoint document library structure
                document_library = f"{folder_url.split('/', 1)[-1]}"
                folder_path = f"{document_library}/{folder_name}"

                # Read file into memory (Prevents closed file issue)
                with open(file_path, "rb") as file:
                    file_content = file.read()  

                # Get SharePoint folder reference
                target_folder = client.web.get_folder_by_server_relative_url(folder_url)

                # Upload file using byte content
                target_folder.upload_file(file_name, file_content)
                
                # Execute request
                client.execute_query()
                orchestrator_connection.log_info(f"✅ Successfully uploaded: {file_name} to {folder_path}")

            except Exception as e:
                orchestrator_connection.log_info(f"❌ Error uploading file: {str(e)}")



# ---- Henter Sagsnummer og Sagsbeskrivelse ---- 
TransactionID = str(uuid.uuid4())
EndDate = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%dT00:00:00")
#FromDate = (datetime.now() - timedelta(days=14)).strftime("%Y-%m-%dT00:00:00")
FromDate = orchestrator_connection.get_constant('NovaEmailExtrator_Timestamp').value
print(FromDate)

payload = {
    "common": {"transactionId": TransactionID},
    "paging": {"startRow": 1, "numberOfRows": 500},
"state": {
    "progressState": "Afsluttet"
},
"states": {
  "startFromDate": FromDate,
  "endFromDate": EndDate,
  "states": [
    {
      "progressState": "Afsluttet"
    }
  ]
},

"buildingCase":{
    "buildingCaseAttributes":{
        "buildingCaseClassName":"Forhåndsdialog"
}
},

  "caseGetOutput": {
    "caseAttributes": {
      "userFriendlyCaseNumber": True,
      "title": True,
    },
    "buildingCase": {
      "buildingCaseAttributes": {
        "buildingCaseClassName": True,
        "buildingCaseClassId": True,
        "bbrCaseId": True,
  }}}}



# Define headers
headers = {
    "Authorization": f"Bearer {KMD_access_token}",
    "Content-Type": "application/json"
}

REMOVE_TITLE_REGEX = re.compile(
    r"(Fejloprettet|Afsluttet\s+mangler\s+fuldmagt)",
    re.IGNORECASE
)

# Define the API endpoint
url = f"{KMDNovaURL}/Case/GetList?api-version=2.0-Case"
# Make the HTTP request
try:
    response = requests.put(url, headers=headers, json=payload)
    response.raise_for_status()  # Raise an error for non-2xx responses
    data = response.json()
    print("Success:", response.status_code)
    
    # Extract and print number of rows
    number_of_rows = data.get("pagingInformation", {}).get("numberOfRows", 0)
    print("Number of Rows:", number_of_rows)

        # Initialize an empty list to store queue items
    Cases = []

    for case in data.get("cases", []):
        case_uuid = case.get("common", {}).get("uuid")
        case_number = case.get("caseAttributes", {}).get("userFriendlyCaseNumber", "Unknown")
        case_title = case.get("caseAttributes", {}).get("title", "")

        # Skip unwanted titles (case-insensitive)
        if case_title and REMOVE_TITLE_REGEX.search(case_title):
            print(f"Removing casenumber {case_number} - because of title: {case_title}")
            continue

        if case_uuid:
            Cases.append({
                "CaseUuid": case_uuid,
                "CaseNumber": case_number,
                "CaseTitle": case_title
            })
except requests.exceptions.RequestException as e:
    print("Request Failed:", e)




#EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
EMAIL_REGEX = re.compile(r"[^\s@]+@[^\s@]+\.[^\s@]+",re.UNICODE)
all_emails = []

# Now iterate through the casesUuid's and get the emails: 

for case in Cases:
    case_uuid = case["CaseUuid"]
    transaction_id = str(uuid.uuid4())
    email_found = False  # <-- IMPORTANT
    payload_case_party = {
        "common": {
            "transactionId": transaction_id,
            "uuid": case_uuid
        },
        "paging": {
            "startRow": 1,
            "numberOfRows": 500
        },
        "caseParty": {
            "partyRole": "IND",
            "partyRoleName": "Indsender"
        },
        "caseGetOutput": {
            "caseAttributes": {
                "userFriendlyCaseNumber": True,
                "title": True,
                "caseDate": True
            },
            "caseParty": {
                "partyRole": True,
                "partyRoleName": True,
                "name": True,
                "participantRole": True,
                "participantRemark": True,
                "participantContactInformation": True
            }
        }
    }

    try:
        response = requests.put(url, headers=headers, json=payload_case_party)
        response.raise_for_status()
        data = response.json()

        for case_item in data.get("cases", []):
            for party in case_item.get("caseParties", []):
                if party.get("partyRole") == "IND":
                    contact_info = party.get("participantContactInformation", "")

                    emails = EMAIL_REGEX.findall(contact_info)

                    if emails:
                        email = emails[0]  # take only the first email

                        print(f"CaseNumber: {case.get('CaseNumber')} | Email: {email}")

                        all_emails.append({
                            "CaseUuid": case_uuid,
                            "CaseNumber": case.get("CaseNumber"),
                            "Email": email
                        })

                        email_found = True
                        break  # stop iterating parties

            if email_found:
                break  # stop iterating cases (defensive)

        # THIS goes here — outside both loops
        if not email_found:
            print(f"CaseNumber: {case.get('CaseNumber')} | Email NOT FOUND")

    except requests.exceptions.RequestException as e:
        print(f"Failed for case {case.get('CaseNumber')} ({case_uuid}): {e}")

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Indsender emails"

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Headers
headers = ["Case number", "Email"]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
for row_idx, item in enumerate(all_emails, start=2):
    ws.cell(row=row_idx, column=1, value=item["CaseNumber"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=item["Email"]).border = thin_border

# Column widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 35

# Freeze header row
ws.freeze_panes = "A2"

# Save file
output_path = "Indsender_emails.xlsx"
wb.save(output_path)

print(f"Excel file created: {output_path}")

orchestrator_connection.log_info('Overfører excelfil til sharepoint')
file_url = f'{SharepointUrl}/Teams/sec-lukket1752/Delte Dokumenter'

certification = orchestrator_connection.get_credential("SharePointCert")
api = orchestrator_connection.get_credential("SharePointAPI")

tenant = api.username
client_id = api.password
thumbprint = certification.username
cert_path = certification.password

client = sharepoint_client(tenant, client_id, thumbprint, cert_path, f'{SharepointUrl}/Teams/tea-teamsite11160/', orchestrator_connection)

upload_to_sharepoint(client= client, folder_name = 'Delte Dokumenter', file_path=output_path, folder_url= '/Teams/tea-teamsite11160/Delte Dokumenter')
orchestrator_connection.log_info(f'Uploaded to {file_url}')

#Opdaterer timestamp: 
orchestrator_connection.update_constant("NovaEmailExtrator_Timestamp",datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))
